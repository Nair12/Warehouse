from django import forms
from django.contrib import admin, messages
from django.contrib.auth.models import Group
from django.db import transaction
from django.core.exceptions import ValidationError
from django.db import models
from django.db.models import Q
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.safestring import mark_safe

from openpyxl import load_workbook

from products.models import Inventory, Product
from warehouses.models import Warehouse

from .models import Trading, TradingItem, TradingAuditLog


try:
    admin.site.unregister(Group)
except admin.sites.NotRegistered:
    pass


class TradingExcelImportForm(forms.Form):
    name = forms.CharField(label="Название сделки", max_length=255)

    trade_type = forms.ChoiceField(
        label="Тип сделки",
        choices=Trading.TradeType.choices,
    )

    comment = forms.CharField(
        label="Комментарий",
        required=False,
        widget=forms.Textarea(attrs={"rows": 4}),
    )

    excel_file = forms.FileField(
        label="Excel файл",
        help_text="Колонки: name, description, warehouse, requested_quantity, fulfilled_quantity",
    )


class TradingItemInline(admin.TabularInline):
    model = TradingItem
    extra = 1
    raw_id_fields = ("product", "warehouse")
    fields = (
        "product",
        "warehouse",
        "quantity",
        "requested_quantity",
        "fulfilled_quantity",
        "quantity_before",
        "quantity_after",
    )


def apply_stock_change_for_admin(trading, item, old_item=None, delete=False):
    """Apply stock impact for one TradingItem edit made from Django admin."""
    if delete:
        product = old_item.product
        warehouse = old_item.warehouse
        quantity = old_item.fulfilled_quantity or 0

        if quantity <= 0:
            return

        inventory, _ = Inventory.objects.select_for_update().get_or_create(
            product=product,
            warehouse=warehouse,
            defaults={"quantity": 0},
        )

        quantity_before = inventory.quantity

        if trading.trade_type == Trading.TradeType.PURCHASE:
            if quantity > inventory.quantity:
                raise ValidationError(
                    f"Нельзя удалить позицию {product}: на складе {inventory.quantity}, нужно откатить {quantity}."
                )
            inventory.quantity -= quantity
        else:
            inventory.quantity += quantity

        inventory.save(update_fields=["quantity", "updated_at"])
        return

    new_quantity = item.fulfilled_quantity or 0

    if old_item is None:
        delta = new_quantity
        product = item.product
        warehouse = item.warehouse
    else:
        old_product_id = old_item.product_id
        old_warehouse_id = old_item.warehouse_id
        new_product_id = item.product_id
        new_warehouse_id = item.warehouse_id

        if old_product_id != new_product_id or old_warehouse_id != new_warehouse_id:
            apply_stock_change_for_admin(trading, item=None, old_item=old_item, delete=True)
            delta = new_quantity
            product = item.product
            warehouse = item.warehouse
        else:
            delta = new_quantity - (old_item.fulfilled_quantity or 0)
            product = item.product
            warehouse = item.warehouse

    inventory, _ = Inventory.objects.select_for_update().get_or_create(
        product=product,
        warehouse=warehouse,
        defaults={"quantity": 0},
    )

    quantity_before = inventory.quantity

    if trading.trade_type == Trading.TradeType.PURCHASE:
        quantity_after = quantity_before + delta
        if quantity_after < 0:
            raise ValidationError(
                f"Нельзя уменьшить позицию {product}: на складе {quantity_before}, нужно откатить {-delta}."
            )
    else:
        quantity_after = quantity_before - delta
        if quantity_after < 0:
            raise ValidationError(
                f"Недостаточно товара на складе для {product}: доступно {quantity_before}, нужно списать {delta}."
            )

    inventory.quantity = quantity_after
    inventory.save(update_fields=["quantity", "updated_at"])

    item.quantity = new_quantity
    item.quantity_before = quantity_before
    item.quantity_after = quantity_after


def refresh_trading_totals_from_items(trading):
    first_item = trading.items.order_by("created_at", "id").first()
    total_fulfilled = trading.items.aggregate(total=models.Sum("fulfilled_quantity"))["total"] or 0
    has_pending_items = trading.items.filter(
        fulfilled_quantity__lt=models.F("requested_quantity")
    ).exists()

    trading.quantity = total_fulfilled
    trading.status = Trading.Status.PENDING if has_pending_items else Trading.Status.COMPLETED

    if first_item:
        trading.product = first_item.product
        trading.warehouse = first_item.warehouse
        trading.quantity_before = first_item.quantity_before
        trading.quantity_after = first_item.quantity_after
    else:
        trading.product = None
        trading.warehouse = None
        trading.quantity_before = 0
        trading.quantity_after = 0

    trading.save(update_fields=[
        "quantity",
        "status",
        "product",
        "warehouse",
        "quantity_before",
        "quantity_after",
        "updated_at",
    ])


@admin.register(Trading)
class TradingAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "name",
        "user",
        "trade_type",
        "product",
        "warehouse",
        "quantity",
        "quantity_before",
        "quantity_after",
    )

    list_filter = (
        "trade_type",
        "warehouse",
        "user",
        "created_at",
    )

    search_fields = (
        "name",
        "product__name",
        "user__username",
        "warehouse__city",
    )

    ordering = ("-created_at",)
    inlines = [TradingItemInline]
    change_list_template = "admin/trading/change_list.html"

    def has_change_permission(self, request, obj=None):
        if not super().has_change_permission(request, obj):
            return False
        return getattr(request.user, "role", None) == "admin" or request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        if not super().has_delete_permission(request, obj):
            return False
        return getattr(request.user, "role", None) == "admin" or request.user.is_superuser

    def save_formset(self, request, form, formset, change):
        if formset.model != TradingItem:
            return super().save_formset(request, form, formset, change)

        trading = form.instance
        before_data = build_trading_after_data(trading) if trading.pk else {}

        try:
            with transaction.atomic():
                old_items = {
                    item.pk: item
                    for item in TradingItem.objects.select_for_update().filter(
                        trading=trading,
                    )
                }

                instances = formset.save(commit=False)

                for deleted_item in getattr(formset, "deleted_objects", []):
                    old_item = old_items.get(deleted_item.pk)
                    if old_item:
                        apply_stock_change_for_admin(
                            trading=trading,
                            item=None,
                            old_item=old_item,
                            delete=True,
                        )
                        old_item.delete()

                for instance in instances:
                    instance.trading = trading

                    if instance.fulfilled_quantity > instance.requested_quantity:
                        raise ValidationError(
                            f"В позиции {instance.product} выполнено больше, чем заказано."
                        )

                    old_item = old_items.get(instance.pk) if instance.pk else None

                    apply_stock_change_for_admin(
                        trading=trading,
                        item=instance,
                        old_item=old_item,
                        delete=False,
                    )
                    instance.save()

                formset.save_m2m()
                refresh_trading_totals_from_items(trading)

                after_data = build_trading_after_data(trading)

                if before_data != after_data:
                    TradingAuditLog.objects.create(
                        trading=trading,
                        trading_id_snapshot=trading.id,
                        user=request.user,
                        action=TradingAuditLog.Action.UPDATED,
                        description="Сделка отредактирована через Django admin. Остатки склада пересчитаны.",
                        before_data=before_data,
                        after_data=after_data,
                    )

        except ValidationError as error:
            messages.error(request, error.messages[0] if hasattr(error, "messages") else str(error))
            raise

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "history-page/",
                self.admin_site.admin_view(self.history_page_view),
                name="trading_trading_history_page",
            ),
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel_view),
                name="trading_trading_import_excel",
            ),
        ]
        return custom_urls + urls

    def history_page_view(self, request):
        tradings = Trading.objects.select_related(
            "product",
            "warehouse",
            "user",
        ).order_by("-created_at")

        context = dict(
            self.admin_site.each_context(request),
            title="История операций по складу",
            tradings=tradings,
        )

        return TemplateResponse(
            request,
            "admin/trading/history_page.html",
            context,
        )

    def import_excel_view(self, request):
        if request.method == "POST":
            form = TradingExcelImportForm(request.POST, request.FILES)

            if form.is_valid():
                try:
                    trading = create_trading_from_excel(
                        excel_file=request.FILES["excel_file"],
                        deal_name=form.cleaned_data["name"],
                        trade_type=form.cleaned_data["trade_type"],
                        comment=form.cleaned_data["comment"],
                        user=request.user,
                    )

                    messages.success(
                        request,
                        f"Сделка создана: {trading.name}. Позиций: {trading.items.count()}.",
                    )
                    return redirect(
                        reverse("admin:trading_trading_change", args=[trading.id])
                    )

                except ValueError as error:
                    messages.error(request, str(error))

                except Exception as error:
                    messages.error(request, f"Ошибка импорта сделки: {error}")
        else:
            form = TradingExcelImportForm()

        context = dict(
            self.admin_site.each_context(request),
            title="Импорт сделки из Excel",
            form=form,
        )

        return TemplateResponse(
            request,
            "admin/trading/import_excel.html",
            context,
        )


@admin.register(TradingItem)
class TradingItemAdmin(admin.ModelAdmin):
    list_display = (
        "trading",
        "product",
        "warehouse",
        "quantity",
        "requested_quantity",
        "fulfilled_quantity",
        "quantity_before",
        "quantity_after",
        "created_at",
    )

    list_filter = (
        "warehouse",
        "created_at",
    )

    search_fields = (
        "trading__name",
        "product__name",
        "warehouse__city",
    )

    ordering = ("-created_at",)


@admin.register(TradingAuditLog)
class TradingAuditLogAdmin(admin.ModelAdmin):
    list_display = (
        "trading_id_snapshot",
        "deal_name_display",
        "last_action_display",
        "user",
        "created_at",
    )

    list_filter = (
        "action",
        "created_at",
        "user",
    )

    search_fields = (
        "description",
        "trading_id_snapshot",
        "user__username",
    )

    ordering = ("-created_at",)

    readonly_fields = (
        "deal_header",
        "deal_history",
    )

    fields = (
        "deal_header",
        "deal_history",
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        queryset = super().get_queryset(request)

        latest_ids = []
        seen_deals = set()

        for log in queryset.order_by("-created_at", "-id"):
            key = log.trading_id_snapshot or log.trading_id

            if key in seen_deals:
                continue

            seen_deals.add(key)
            latest_ids.append(log.id)

        return queryset.filter(id__in=latest_ids)

    def get_search_results(self, request, queryset, search_term):
        search_term = (search_term or "").strip().lower()

        if not search_term:
            return queryset, False

        matched_deal_ids = set()

        for log in TradingAuditLog.objects.select_related("user"):
            deal_id = log.trading_id_snapshot or log.trading_id

            before = log.before_data or {}
            after = log.after_data or {}

            if (
                search_term in str(before.get("name", "")).lower()
                or search_term in str(after.get("name", "")).lower()
                or search_term in str(log.description or "").lower()
            ):
                matched_deal_ids.add(deal_id)

        if matched_deal_ids:
            queryset = self.get_queryset(request).filter(
                Q(trading_id_snapshot__in=matched_deal_ids) |
                Q(trading_id__in=matched_deal_ids)
            )
        else:
            queryset = self.get_queryset(request).none()

        return queryset, False

    def deal_name_display(self, obj):
        data = obj.after_data or obj.before_data or {}
        return data.get("name") or "Без названия"

    deal_name_display.short_description = "Название сделки"

    def last_action_display(self, obj):
        return obj.get_action_display()

    last_action_display.short_description = "Последнее действие"

    def get_all_logs_for_deal(self, obj):
        deal_id = obj.trading_id_snapshot or obj.trading_id

        return TradingAuditLog.objects.filter(
            Q(trading_id_snapshot=deal_id) | Q(trading_id=deal_id)
        ).select_related("user").order_by("created_at", "id")

    def deal_header(self, obj):
        data = obj.after_data or obj.before_data or {}
        deal_name = data.get("name") or "Без названия"
        deal_id = obj.trading_id_snapshot or obj.trading_id
        trade_type = data.get("trade_type")

        if trade_type == "purchase":
            trade_label = "Покупка"
            trade_color = "#2ecc71"
            trade_icon = "+"
        elif trade_type == "sell":
            trade_label = "Продажа"
            trade_color = "#e74c3c"
            trade_icon = "-"
        else:
            trade_label = "-"
            trade_color = "#888"
            trade_icon = "•"

        logs_count = TradingAuditLog.objects.filter(
            Q(trading_id_snapshot=deal_id) | Q(trading_id=deal_id)
        ).count()

        return mark_safe(f'''
        <style>
            .form-row.field-deal_header > div > label,
            .form-row.field-deal_history > div > label,
            .form-row.field-deal_header label,
            .form-row.field-deal_history label {{
                display: none !important;
            }}

            .form-row.field-deal_header,
            .form-row.field-deal_history {{
                padding-left: 0 !important;
            }}

            .form-row.field-deal_header > div,
            .form-row.field-deal_history > div {{
                margin-left: 0 !important;
                width: 100% !important;
            }}
        </style>

        <div style="background:#1e1e1e;border:1px solid #333;border-radius:14px;padding:18px;margin-bottom:18px;margin-left:0;max-width:900px;width:100%;box-sizing:border-box;">
            <div style="font-size:13px;color:#aaa;margin-bottom:6px;">История сделки</div>
            <div style="font-size:26px;font-weight:800;color:white;">{deal_name}</div>
            <div style="margin-top:10px;">
                <span style="background:{trade_color};color:white;padding:5px 12px;border-radius:999px;font-size:13px;font-weight:700;">
                    {trade_icon} {trade_label}
                </span>
            </div>
            <div style="margin-top:14px;color:#aaa;">
                ID: <b style="color:white;">#{deal_id}</b> |
                Действий: <b style="color:white;">{logs_count}</b>
            </div>
        </div>
        ''')

    deal_header.short_description = "Сделка"

    def deal_history(self, obj):
        logs = self.get_all_logs_for_deal(obj)

        html = '<div style="display:flex;flex-direction:column;gap:12px;margin-left:0;max-width:950px;width:100%;">'

        for log in logs:
            html += self.render_log_card(log)

        html += "</div>"
        return mark_safe(html)

    def render_log_card(self, log):
        colors = {
            "created": "#2ecc71",
            "fulfilled": "#3498db",
            "updated": "#f1c40f",
            "deleted": "#e74c3c",
            "rollback": "#8a4fb8",
            "Создание": "#2ecc71",
            "Дополнение": "#3498db",
            "Редактирование": "#f1c40f",
            "Удаление": "#e74c3c",
            "Откат склада": "#8a4fb8",
        }

        color = colors.get(log.action) or colors.get(log.get_action_display(), "#4f7c8a")

        return f'''
        <div style="background:#1e1e1e;border-left:6px solid {color};border-radius:12px;padding:14px;">
            <div style="display:flex;justify-content:space-between;">
                <div>
                    <span style="color:{color};font-weight:800;">{log.get_action_display()}</span>
                    <div style="margin-top:5px;">{log.description}</div>
                </div>
                <div style="text-align:right;">
                    <div style="color:#aaa;font-size:12px;">
                        {log.created_at.strftime('%d.%m.%Y %H:%M:%S')}
                    </div>
                    <div style="color:#fff;font-weight:700;">
                        User: {log.user or '-'}
                    </div>
                </div>
            </div>
            <div style="margin-top:10px;">
                {self.render_changes(log)}
            </div>
        </div>
        '''

    def render_changes(self, log):
        before = log.before_data or {}
        after = log.after_data or {}

        before_items = before.get("items", [])
        after_items = after.get("items", [])

        item_fields = {
            "requested_quantity": "Заказано",
            "fulfilled_quantity": "Выполнено",
            "quantity_after": "Остаток на складе",
        }

        grouped = {}

        for i in range(max(len(before_items), len(after_items))):
            old_item = before_items[i] if i < len(before_items) else None
            new_item = after_items[i] if i < len(after_items) else None

            if not old_item or not new_item:
                continue

            product = new_item.get("product") or "Товар"

            grouped.setdefault(product, [])

            for key, label in item_fields.items():
                old = old_item.get(key)
                new = new_item.get(key)

                if old != new:
                    grouped[product].append((label, old, new))

        grouped = {
            product: changes
            for product, changes in grouped.items()
            if changes
        }

        if not grouped:
            return '<div style="color:#aaa;">Нет изменений</div>'

        html = '<div style="display:flex;flex-direction:column;gap:10px;">'

        for product, changes in grouped.items():
            html += f'''
            <div style="background:#151515;border:1px solid #2a2a2a;border-radius:10px;padding:14px;width:100%;box-sizing:border-box;">
                <div style="font-weight:800;margin-bottom:8px;color:#fff;">
                    Product: {product}
                </div>
                <table style="width:100%;border-collapse:collapse;table-layout:fixed;">
            '''

            for label, old, new in changes:
                html += f'''
                <tr>
                    <td style="padding:8px 10px;color:#ddd;width:40%;font-size:14px;">{label}</td>
                    <td style="padding:8px 10px;width:30%;font-size:14px;">
                        <span style="color:#aaa;">было:</span>
                        <span style="color:#e74c3c;font-weight:600;"> {old}</span>
                    </td>
                    <td style="padding:8px 10px;width:30%;font-size:14px;">
                        <span style="color:#aaa;">стало:</span>
                        <span style="color:#2ecc71;font-weight:700;"> {new}</span>
                    </td>
                </tr>
                '''

            html += "</table></div>"

        html += "</div>"
        return html


def normalize_header(value):
    if value is None:
        return ""

    value = str(value).strip().lower()
    value = value.replace("\n", " ").replace("\r", " ")
    return " ".join(value.split())


def find_header(headers, names):
    normalized_names = [normalize_header(name) for name in names]

    for name in normalized_names:
        if name in headers:
            return headers.index(name) + 1

    for index, header in enumerate(headers):
        for name in normalized_names:
            if name and name in header:
                return index + 1

    return None


def clean_text(value):
    return " ".join(str(value or "").strip().split())

def parse_positive_int(value, field_name, row_number, errors):
    if value is None or str(value).strip() == "":
        errors.append(f"Строка {row_number}: поле '{field_name}' пустое.")
        return None

    raw = str(value).strip().replace(" ", "").replace(",", ".")

    try:
        number = float(raw)
    except ValueError:
        errors.append(f"Строка {row_number}: поле '{field_name}' не число: {value}.")
        return None

    if number < 0:
        errors.append(f"Строка {row_number}: поле '{field_name}' не может быть отрицательным.")
        return None

    if number != int(number):
        errors.append(f"Строка {row_number}: поле '{field_name}' должно быть целым: {value}.")
        return None

    return int(number)


def get_product_for_import(raw_code, row_number, errors):
    code = clean_text(raw_code)

    if not code:
        errors.append(f"Строка {row_number}: пустой код товара.")
        return None

    lookup_code = code.replace(" ", "").casefold()

    for product in Product.objects.only("id", "name"):
        product_code = (product.name or "").replace(" ", "").casefold()

        if product_code == lookup_code:
            return product

    errors.append(
        f'Строка {row_number}: товар с кодом "{code}" не найден в базе.'
    )
    return None

def get_warehouse_for_import(raw_warehouse, row_number, errors):
    warehouse_name = clean_text(raw_warehouse)

    if not warehouse_name:
        errors.append(f"Строка {row_number}: пустой склад.")
        return None

    warehouse = Warehouse.objects.filter(city__iexact=warehouse_name).first()

    if not warehouse:
        errors.append(f"Строка {row_number}: склад не найден: {warehouse_name}.")
        return None

    return warehouse


def build_trading_after_data(trading):
    items = []

    for item in trading.items.select_related("product", "warehouse").all():
        items.append({
            "product": item.product.name if item.product else "",
            "warehouse": item.warehouse.city if item.warehouse else "",
            "quantity": item.quantity,
            "requested_quantity": item.requested_quantity,
            "fulfilled_quantity": item.fulfilled_quantity,
            "quantity_before": item.quantity_before,
            "quantity_after": item.quantity_after,
        })

    return {
        "id": trading.id,
        "name": trading.name,
        "trade_type": trading.trade_type,
        "status": trading.status,
        "comment": trading.comment,
        "items": items,
    }


def create_trading_from_excel(excel_file, deal_name, trade_type, comment, user):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active

    headers = [normalize_header(cell.value) for cell in sheet[1]]

    code_index = find_header(headers, [
        "name", "код", "код товара", "чертежный номер", "чертёжный номер",
        "part no", "part number", "pn",
    ])
    description_index = find_header(headers, [
        "description", "описание", "название", "наименование",
    ])
    warehouse_index = find_header(headers, [
        "warehouse", "склад", "город",
    ])
    requested_index = find_header(headers, [
        "requested_quantity", "requested", "заказано", "нужно", "количество заказано",
    ])
    fulfilled_index = find_header(headers, [
        "fulfilled_quantity", "fulfilled", "получено", "отдали", "выполнено",
        "количество получено", "quantity", "qty", "количество",
    ])

    errors = []

    if not code_index:
        errors.append("Не найдена колонка товара: name / код / part number.")

    if not description_index:
        errors.append("Не найдена колонка описания: description / описание.")

    if not warehouse_index:
        errors.append("Не найдена колонка склада: warehouse / склад.")

    if not requested_index:
        errors.append("Не найдена колонка заказанного количества: requested_quantity / заказано.")

    if not fulfilled_index:
        errors.append("Не найдена колонка полученного количества: fulfilled_quantity / получено.")

    if errors:
        raise ValueError("\n".join(errors))

    parsed_rows = []

    for row_number in range(2, sheet.max_row + 1):
        raw_code = sheet.cell(row=row_number, column=code_index).value
        raw_description = sheet.cell(row=row_number, column=description_index).value
        raw_warehouse = sheet.cell(row=row_number, column=warehouse_index).value
        raw_requested = sheet.cell(row=row_number, column=requested_index).value
        raw_fulfilled = sheet.cell(row=row_number, column=fulfilled_index).value

        if (
            not clean_text(raw_code)
            and not clean_text(raw_description)
            and not clean_text(raw_warehouse)
            and not clean_text(raw_requested)
            and not clean_text(raw_fulfilled)
        ):
            continue

        product = get_product_for_import(raw_code, row_number, errors)
        warehouse = get_warehouse_for_import(raw_warehouse, row_number, errors)
        requested_quantity = parse_positive_int(raw_requested, "requested_quantity", row_number, errors)
        fulfilled_quantity = parse_positive_int(raw_fulfilled, "fulfilled_quantity", row_number, errors)

        if (
            requested_quantity is not None
            and fulfilled_quantity is not None
            and fulfilled_quantity > requested_quantity
        ):
            errors.append(
                f"Строка {row_number}: получено/отдано больше, чем заказано "
                f"({fulfilled_quantity} > {requested_quantity})."
            )

        if product and warehouse and requested_quantity is not None and fulfilled_quantity is not None:
            parsed_rows.append({
                "row_number": row_number,
                "product": product,
                "warehouse": warehouse,
                "requested_quantity": requested_quantity,
                "fulfilled_quantity": fulfilled_quantity,
            })

    if not parsed_rows:
        errors.append("В Excel нет строк для импорта.")

    if errors:
        raise ValueError("Импорт остановлен. Ничего не создано.\n\n" + "\n\n".join(errors[:50]))

    with transaction.atomic():
        product_ids = [row["product"].id for row in parsed_rows]
        warehouse_ids = [row["warehouse"].id for row in parsed_rows]

        inventories = Inventory.objects.select_for_update().filter(
            product_id__in=product_ids,
            warehouse_id__in=warehouse_ids,
        )

        inventory_by_key = {
            (inventory.product_id, inventory.warehouse_id): inventory
            for inventory in inventories
        }

        stock_errors = []

        for row in parsed_rows:
            product = row["product"]
            warehouse = row["warehouse"]
            fulfilled_quantity = row["fulfilled_quantity"]

            inventory = inventory_by_key.get((product.id, warehouse.id))
            current_quantity = inventory.quantity if inventory else 0

            if trade_type == Trading.TradeType.SELL and current_quantity < fulfilled_quantity:
                stock_errors.append(
                    "Строка {row}: недостаточно товара для продажи.\n"
                    "Товар: {product}\n"
                    "Склад: {warehouse}\n"
                    "Нужно списать: {need}\n"
                    "Есть на складе: {current}".format(
                        row=row["row_number"],
                        product=product.name,
                        warehouse=warehouse.city,
                        need=fulfilled_quantity,
                        current=current_quantity,
                    )
                )

        if stock_errors:
            raise ValueError("Импорт остановлен. Ничего не создано.\n\n" + "\n\n".join(stock_errors[:50]))

        first_row = parsed_rows[0]

        trading = Trading.objects.create(
            name=deal_name,
            comment=comment,
            trade_type=trade_type,
            status=Trading.Status.COMPLETED,
            user=user,
            product=first_row["product"],
            warehouse=first_row["warehouse"],
            quantity=sum(row["fulfilled_quantity"] for row in parsed_rows),
            quantity_before=0,
            quantity_after=0,
        )

        for row in parsed_rows:
            product = row["product"]
            warehouse = row["warehouse"]
            requested_quantity = row["requested_quantity"]
            fulfilled_quantity = row["fulfilled_quantity"]

            inventory = inventory_by_key.get((product.id, warehouse.id))

            if not inventory:
                inventory = Inventory.objects.create(
                    product=product,
                    warehouse=warehouse,
                    quantity=0,
                )
                inventory_by_key[(product.id, warehouse.id)] = inventory

            quantity_before = inventory.quantity

            if trade_type == Trading.TradeType.PURCHASE:
                quantity_after = quantity_before + fulfilled_quantity
            else:
                quantity_after = quantity_before - fulfilled_quantity

            inventory.quantity = quantity_after
            inventory.save(update_fields=["quantity", "updated_at"])

            TradingItem.objects.create(
                trading=trading,
                product=product,
                warehouse=warehouse,
                quantity=fulfilled_quantity,
                requested_quantity=requested_quantity,
                fulfilled_quantity=fulfilled_quantity,
                quantity_before=quantity_before,
                quantity_after=quantity_after,
            )

        first_item = trading.items.order_by("created_at").first()

        if first_item:
            trading.quantity_before = first_item.quantity_before
            trading.quantity_after = first_item.quantity_after
            trading.save(update_fields=["quantity_before", "quantity_after", "updated_at"])

        TradingAuditLog.objects.create(
            trading=trading,
            trading_id_snapshot=trading.id,
            user=user,
            action=TradingAuditLog.Action.CREATED,
            description=f"Сделка создана через импорт Excel. Позиций: {trading.items.count()}.",
            before_data={},
            after_data=build_trading_after_data(trading),
        )

        return trading
