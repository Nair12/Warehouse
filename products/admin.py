import re
from difflib import SequenceMatcher

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html

from openpyxl import Workbook, load_workbook

from warehouses.models import Warehouse

from .models import (
    Inventory,
    Product,
    ProductImportBatch,
    ProductImportRow,
    normalize_product_code,
)


class ProductImportExcelForm(forms.Form):
    """
    Простая форма для администратора.

    Сейчас админ должен только выбрать Excel.
    Остальные настройки оставляем внутри системы, чтобы не перегружать интерфейс:
    - импортируем только товары;
    - остатки пока не трогаем;
    - режим количества заранее задан для будущего расширения.
    """

    excel_file = forms.FileField(
        label="Excel файл",
        help_text="Загрузите Excel со списком товаров. Система сама попробует найти нужные колонки.",
    )

    import_type = forms.ChoiceField(
        label="Тип импорта",
        choices=ProductImportBatch.IMPORT_TYPE_CHOICES,
        initial=ProductImportBatch.IMPORT_TYPE_PRODUCTS_ONLY,
        required=False,
        widget=forms.HiddenInput,
    )

    target_warehouse = forms.ModelChoiceField(
        label="Склад",
        queryset=Warehouse.objects.all().order_by("city"),
        required=False,
        empty_label="Выберите склад",
        widget=forms.Select(attrs={"class": "form-control"}),
    )

    quantity_mode = forms.ChoiceField(
        label="Как обновлять остатки",
        choices=ProductImportBatch.QUANTITY_MODE_CHOICES,
        initial=ProductImportBatch.QUANTITY_MODE_REPLACE,
        required=False,
        widget=forms.HiddenInput,
    )

    def clean(self):
        cleaned_data = super().clean()

        import_type = cleaned_data.get("import_type") or ProductImportBatch.IMPORT_TYPE_PRODUCTS_ONLY
        quantity_mode = cleaned_data.get("quantity_mode") or ProductImportBatch.QUANTITY_MODE_REPLACE
        target_warehouse = cleaned_data.get("target_warehouse")

        cleaned_data["import_type"] = import_type
        cleaned_data["quantity_mode"] = quantity_mode
        cleaned_data["target_warehouse"] = target_warehouse

        if import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK and not target_warehouse:
            raise forms.ValidationError("Для импорта остатков нужно выбрать склад.")

        return cleaned_data

def normalize_unit(value):
    if value is None:
        return Product.UNIT_PIECE

    value = str(value).strip().lower()

    if value in ["", "шт", "штук", "pcs", "pc", "piece", "pieces", "ед", "ед."]:
        return Product.UNIT_PIECE

    if value in ["кг", "kg", "kgs", "килограмм", "килограммы"]:
        return Product.UNIT_KG

    return Product.UNIT_PIECE


def parse_quantity(value):
    if value is None:
        return None, "Количество пустое."

    raw = str(value).strip()

    if raw == "":
        return None, "Количество пустое."

    raw = raw.replace(",", ".")
    raw = raw.replace(" ", "")

    try:
        number = float(raw)
    except ValueError:
        return None, f"Не удалось прочитать количество: {value}"

    if number < 0:
        return None, "Количество не может быть отрицательным."

    if number != int(number):
        return None, f"Количество должно быть целым числом: {value}"

    return int(number), ""


def import_description_looks_like_code(raw_code, raw_name):
    """
    Проверяет, похоже ли название из Excel просто на код.
    Если да — не используем его как нормальное описание товара.
    """
    raw_code_text = str(raw_code or "").strip()
    raw_name_text = str(raw_name or "").strip()

    if not raw_name_text:
        return True

    if raw_code_text and raw_code_text.upper() == raw_name_text.upper():
        return True

    normalized_code = normalize_product_code(raw_code_text)
    normalized_name = normalize_product_code(raw_name_text)

    if normalized_code and normalized_code == normalized_name:
        return True

    return False


def build_import_product_description(raw_code, raw_name):
    """
    Описание для НОВОГО товара.

    Важно:
    - существующим товарам description не меняем вообще;
    - если в Excel вместо описания повторяется код — ставим нейтральное описание;
    - если описание есть — используем его.
    """
    raw_name_text = str(raw_name or "").strip()

    if import_description_looks_like_code(raw_code, raw_name):
        return "Деталь / Part"

    return raw_name_text or "Деталь / Part"


def text_similarity(first, second):
    first = str(first or "").strip().upper()
    second = str(second or "").strip().upper()

    if not first or not second:
        return 0

    return int(SequenceMatcher(None, first, second).ratio() * 100)


def code_similarity(first, second):
    first = normalize_product_code(first)
    second = normalize_product_code(second)

    if not first or not second:
        return 0

    if first == second:
        return 100

    if first in second or second in first:
        return 92

    return int(SequenceMatcher(None, first, second).ratio() * 100)


def is_exact_code_match(first, second):
    first = str(first or "").strip().upper()
    second = str(second or "").strip().upper()
    return bool(first and second and first == second)


def _code_without_separators(value):
    if not value:
        return ""

    value = str(value).strip().upper()
    return re.sub(r"[^A-ZА-ЯЁ0-9]", "", value)


def _code_without_all_zeros(value):
    return _code_without_separators(value).replace("0", "")


def _code_without_leading_zeros(value):
    value = _code_without_separators(value)
    return value.lstrip("0") or "0"


def find_best_product_match(raw_code, raw_name=None):
    """
    Безопасный поиск товара по чертежному номеру.

    Главное правило:
    - точное совпадение после удаления разделителей -> можно импортировать автоматически;
    - совпадение только из-за нулей -> ручная проверка;
    - отличается реальная цифра/буква -> другой товар.
    """
    raw_code_text = str(raw_code or "").strip()

    strict_code = _code_without_separators(raw_code_text)
    without_leading_zeros = _code_without_leading_zeros(raw_code_text)
    without_all_zeros = _code_without_all_zeros(raw_code_text)

    if not strict_code:
        return None, ProductImportRow.STATUS_ERROR, "Нет чертежного номера."

    exact_product = Product.objects.filter(name__iexact=raw_code_text).first()
    if exact_product:
        return exact_product, ProductImportRow.STATUS_EXISTING, "Точное совпадение по чертежному номеру."

    strict_matches = {}

    for product in Product.objects.only("id", "name", "normalized_code").iterator():
        product_strict = _code_without_separators(product.name)

        if product_strict == strict_code:
            strict_matches[product.id] = product

            if len(strict_matches) > 1:
                break

    if len(strict_matches) == 1:
        product = list(strict_matches.values())[0]
        return (
            product,
            ProductImportRow.STATUS_EXISTING,
            (
                "Товар найден по чертежному номеру. "
                "Разделители не считаются отличием."
            ),
        )

    if len(strict_matches) > 1:
        return (
            None,
            ProductImportRow.STATUS_SIMILAR_CODE,
            (
                "Найдено несколько товаров с одинаковым кодом после удаления разделителей. "
                "Нужно проверить вручную."
            ),
        )

    zero_related_matches = {}

    for product in Product.objects.only("id", "name", "normalized_code").iterator():
        product_without_leading_zeros = _code_without_leading_zeros(product.name)
        product_without_all_zeros = _code_without_all_zeros(product.name)

        if (
            product_without_leading_zeros == without_leading_zeros
            or product_without_all_zeros == without_all_zeros
        ):
            zero_related_matches[product.id] = product

            if len(zero_related_matches) > 1:
                break

    if len(zero_related_matches) == 1:
        product = list(zero_related_matches.values())[0]
        return (
            product,
            ProductImportRow.STATUS_SIMILAR_CODE,
            (
                "Код похож на товар в базе только после обработки нулей. "
                "Нужно проверить вручную, чтобы не добавить количество не туда."
            ),
        )

    if len(zero_related_matches) > 1:
        return (
            None,
            ProductImportRow.STATUS_SIMILAR_CODE,
            (
                "После обработки нулей найдено несколько похожих товаров. "
                "Автоматически импортировать опасно, нужна ручная проверка."
            ),
        )

    return None, ProductImportRow.STATUS_NEW, "Товар не найден в базе. Можно создать новый товар."


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ("name", "normalized_code", "id")
    search_fields = ("name", "normalized_code", "description")
    list_filter = ()
    ordering = ("name",)

    change_list_template = "admin/products/product/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel),
                name="products_product_import_excel",
            ),
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel),
                name="products_product_export_excel",
            ),
            path(
                "export-excel-with-stock/",
                self.admin_site.admin_view(self.export_excel_with_stock),
                name="products_product_export_excel_with_stock",
            ),
        ]
        return custom_urls + urls

    def export_excel(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        sheet.append(["name", "normalized_code", "description"])

        products = Product.objects.all().order_by("name")

        for product in products:
            sheet.append([
                product.name,
                product.normalized_code,
                product.description or "",
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="products_export.xlsx"'

        workbook.save(response)
        return response

    def export_excel_with_stock(self, request):
        """
        Экспорт товаров вместе с остатками по каждому складу.

        Старый export_excel не трогаем.
        Этот экспорт нужен, чтобы скачать актуальную базу с Azure
        и безопасно перенести её на локальный сайт для тестов.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products with stock"

        warehouses = list(Warehouse.objects.all().order_by("city"))

        headers = ["name", "normalized_code", "description"]
        headers += [warehouse.city for warehouse in warehouses]
        headers += ["total_quantity"]
        sheet.append(headers)

        products = Product.objects.all().order_by("name")

        inventory_rows = Inventory.objects.select_related("product", "warehouse").filter(
            product__in=products,
            warehouse__in=warehouses,
        )

        quantity_by_product_and_warehouse = {
            (inventory.product_id, inventory.warehouse_id): inventory.quantity
            for inventory in inventory_rows
        }

        for product in products:
            row = [
                product.name,
                product.normalized_code,
                product.description or "",
            ]

            total_quantity = 0

            for warehouse in warehouses:
                quantity = quantity_by_product_and_warehouse.get(
                    (product.id, warehouse.id),
                    0,
                )
                total_quantity += quantity
                row.append(quantity)

            row.append(total_quantity)
            sheet.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="products_export_with_stock.xlsx"'

        workbook.save(response)
        return response

    def import_excel(self, request):
        if request.method == "POST":
            form = ProductImportExcelForm(request.POST, request.FILES)

            if form.is_valid():
                try:
                    batch = create_product_import_batch_from_excel(
                        excel_file=request.FILES["excel_file"],
                        user=request.user,
                        import_type=form.cleaned_data["import_type"],
                        target_warehouse=form.cleaned_data["target_warehouse"],
                        quantity_mode=form.cleaned_data["quantity_mode"],
                    )

                    messages.success(
                        request,
                        (
                            f"Excel проверен. Всего строк: {batch.total_rows}. "
                            f"Уже есть: {batch.existing_count}. "
                            f"Похожие коды: {batch.similar_code_count}. "
                            f"Похожие названия: {batch.similar_name_count}. "
                            f"Дубли в Excel: {batch.duplicate_in_file_count}. "
                            f"Новые: {batch.new_count}. "
                            f"Ошибки: {batch.error_count + batch.quantity_error_count}."
                        ),
                    )

                    return redirect(f"/admin/products/productimportbatch/{batch.id}/change/")

                except Exception as e:
                    messages.error(request, f"Ошибка импорта Excel: {e}")
                    return redirect("..")
        else:
            form = ProductImportExcelForm()

        context = {
            **self.admin_site.each_context(request),
            "title": "Черновой импорт товаров из Excel",
            "form": form,
        }

        return render(request, "admin/products/product/import_excel.html", context)


@admin.register(Inventory)
class InventoryAdmin(admin.ModelAdmin):
    list_display = ("product_name", "warehouse_city", "quantity")
    search_fields = ("product__name", "warehouse__city")
    list_filter = ("warehouse__city",)
    ordering = ("warehouse__city", "product__name")

    @admin.display(description="Товар", ordering="product__name")
    def product_name(self, obj):
        return obj.product.name

    @admin.display(description="Город", ordering="warehouse__city")
    def warehouse_city(self, obj):
        return obj.warehouse.city


class ProductImportRowInline(admin.TabularInline):
    model = ProductImportRow
    extra = 0
    fields = (
        "row_number",
        "raw_code",
        "raw_name",
        "raw_quantity",
        "quantity",
        "detected_status",
        "suggested_product",
        "selected_product",
        "action",
        "note",
        "is_processed",
        "created_product",
        "inventory_updated",
        "inventory_before",
        "inventory_after",
    )
    readonly_fields = (
        "row_number",
        "raw_code",
        "raw_name",
        "raw_quantity",
        "quantity",
        "detected_status",
        "suggested_product",
        "note",
        "is_processed",
        "created_product",
        "inventory_updated",
        "inventory_before",
        "inventory_after",
    )
    autocomplete_fields = ("selected_product",)
    can_delete = False
    show_change_link = True


@admin.register(ProductImportBatch)
class ProductImportBatchAdmin(admin.ModelAdmin):
    list_display = (
        "title",
        "import_type",
        "target_warehouse",
        "status_badge",
        "total_rows",
        "attention_count_display",
        "existing_count",
        "new_count",
        "updated_inventory_count",
        "created_at",
    )
    list_filter = ("status", "import_type", "target_warehouse", "created_at")
    search_fields = ("title",)
    readonly_fields = (
        "status",
        "created_by",
        "total_rows",
        "existing_count",
        "similar_code_count",
        "similar_name_count",
        "duplicate_in_file_count",
        "new_count",
        "error_count",
        "quantity_error_count",
        "created_products_count",
        "updated_inventory_count",
        "skipped_rows_count",
        "linked_rows_count",
        "created_at",
        "updated_at",
    )
    inlines = (ProductImportRowInline,)
    actions = ("recheck_import_rows", "confirm_import_rows", "cancel_import")
    change_list_template = "admin/products/productimportbatch/change_list.html"
    change_form_template = "admin/products/productimportbatch/change_form.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "upload-excel/",
                self.admin_site.admin_view(self.upload_excel),
                name="products_productimportbatch_upload_excel",
            ),
            path(
                "<uuid:batch_id>/confirm/",
                self.admin_site.admin_view(self.confirm_single_import),
                name="products_productimportbatch_confirm_single",
            ),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        imports = ProductImportBatch.objects.all()

        extra_context = extra_context or {}
        extra_context.update({
            "dashboard_total_products": Product.objects.count(),
            "dashboard_total_imports": imports.count(),
            "dashboard_waiting_imports": imports.exclude(
                status__in=[
                    ProductImportBatch.STATUS_IMPORTED,
                    ProductImportBatch.STATUS_CANCELLED,
                ]
            ).count(),
            "dashboard_problem_rows": ProductImportRow.objects.filter(
                detected_status__in=[
                    ProductImportRow.STATUS_SIMILAR_CODE,
                    ProductImportRow.STATUS_SIMILAR_NAME,
                    ProductImportRow.STATUS_DUPLICATE_IN_FILE,
                    ProductImportRow.STATUS_ERROR,
                    ProductImportRow.STATUS_QUANTITY_ERROR,
                ],
                is_processed=False,
            ).count(),
            "dashboard_recent_imports": imports.order_by("-created_at")[:5],
            "upload_form": ProductImportExcelForm(),
        })

        return super().changelist_view(request, extra_context=extra_context)

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = extra_context or {}

        if object_id:
            batch = self.get_object(request, object_id)
            if batch:
                extra_context.update({
                    "attention_rows_count": batch.rows.filter(
                        detected_status__in=[
                            ProductImportRow.STATUS_SIMILAR_CODE,
                            ProductImportRow.STATUS_SIMILAR_NAME,
                            ProductImportRow.STATUS_DUPLICATE_IN_FILE,
                            ProductImportRow.STATUS_ERROR,
                            ProductImportRow.STATUS_QUANTITY_ERROR,
                        ],
                        is_processed=False,
                    ).count(),
                    "existing_rows_count": batch.rows.filter(detected_status=ProductImportRow.STATUS_EXISTING).count(),
                    "new_rows_count": batch.rows.filter(detected_status=ProductImportRow.STATUS_NEW).count(),
                    "processed_rows_count": batch.rows.filter(is_processed=True).count(),
                })

        return super().changeform_view(request, object_id, form_url, extra_context)

    def upload_excel(self, request):
        if request.method != "POST":
            return redirect("..")

        form = ProductImportExcelForm(request.POST, request.FILES)

        if not form.is_valid():
            for error in form.non_field_errors():
                messages.error(request, error)
            messages.error(request, "Проверь выбранный файл, тип импорта и склад.")
            return redirect("..")

        try:
            batch = create_product_import_batch_from_excel(
                excel_file=request.FILES["excel_file"],
                user=request.user,
                import_type=form.cleaned_data["import_type"],
                target_warehouse=form.cleaned_data["target_warehouse"],
                quantity_mode=form.cleaned_data["quantity_mode"],
            )

            messages.success(
                request,
                (
                    f"Excel проверен. Всего строк: {batch.total_rows}. "
                    f"Уже есть: {batch.existing_count}. "
                    f"Требуют проверки: {batch.attention_count}. "
                    f"Новые: {batch.new_count}. "
                    f"Ошибки: {batch.error_count + batch.quantity_error_count}."
                ),
            )

            return redirect(f"../{batch.id}/change/")

        except Exception as e:
            messages.error(request, f"Ошибка загрузки Excel: {e}")
            return redirect("..")

    def confirm_single_import(self, request, batch_id):
        if request.method != "POST":
            return redirect(f"../../{batch_id}/change/")

        batch = ProductImportBatch.objects.filter(id=batch_id).first()

        if not batch:
            messages.error(request, "Импорт не найден.")
            return redirect("../..")

        try:
            result = confirm_product_import_batch(batch)

            messages.success(
                request,
                (
                    f"Импорт завершён. "
                    f"Создано товаров: {result['created']}. "
                    f"Привязано к существующим: {result['linked']}. "
                    f"Обновлено остатков: {result['inventory_updated']}. "
                    f"Пропущено строк: {result['skipped']}."
                ),
            )

        except Exception as e:
            messages.error(request, f"Ошибка подтверждения импорта: {e}")

        return redirect(f"../../{batch_id}/change/")


    @admin.display(description="Статус")
    def status_badge(self, obj):
        colors = {
            ProductImportBatch.STATUS_DRAFT: "#6c757d",
            ProductImportBatch.STATUS_CHECKED: "#0d6efd",
            ProductImportBatch.STATUS_IMPORTED: "#198754",
            ProductImportBatch.STATUS_CANCELLED: "#dc3545",
        }
        color = colors.get(obj.status, "#6c757d")
        return format_html(
            '<span style="padding:4px 8px;border-radius:12px;background:{};color:white;">{}</span>',
            color,
            obj.get_status_display(),
        )

    @admin.display(description="Требуют проверки")
    def attention_count_display(self, obj):
        count = obj.attention_count
        if count == 0:
            return format_html(
            '<span style="color:{};font-weight:700;">{}</span>',
            "#22c55e",
            "0",
        )
        return format_html('<span style="color:#f97316;font-weight:700;">{}</span>', count)

    def save_model(self, request, obj, form, change):
        if not obj.created_by_id:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

    @admin.action(description="Перепроверить строки импорта")
    def recheck_import_rows(self, request, queryset):
        checked = 0

        for batch in queryset:
            recheck_product_import_batch(batch)
            checked += 1

        messages.success(request, f"Перепроверено импортов: {checked}.")

    @admin.action(description="Подтвердить импорт выбранных файлов")
    def confirm_import_rows(self, request, queryset):
        imported_batches = 0
        created_products = 0
        skipped_rows = 0
        linked_rows = 0
        updated_inventory = 0

        for batch in queryset:
            result = confirm_product_import_batch(batch)
            imported_batches += 1
            created_products += result["created"]
            skipped_rows += result["skipped"]
            linked_rows += result["linked"]
            updated_inventory += result["inventory_updated"]

        messages.success(
            request,
            (
                f"Импорт подтверждён. Импортов: {imported_batches}. "
                f"Создано товаров: {created_products}. "
                f"Привязано к существующим: {linked_rows}. "
                f"Обновлено остатков: {updated_inventory}. "
                f"Пропущено строк: {skipped_rows}."
            ),
        )

    @admin.action(description="Отменить выбранные импорты")
    def cancel_import(self, request, queryset):
        updated = queryset.exclude(status=ProductImportBatch.STATUS_IMPORTED).update(
            status=ProductImportBatch.STATUS_CANCELLED
        )
        messages.success(request, f"Отменено импортов: {updated}.")


@admin.register(ProductImportRow)
class ProductImportRowAdmin(admin.ModelAdmin):
    change_list_template = "admin/products/productimportrow/change_list.html"
    list_per_page = 50
    show_full_result_count = False

    list_display = (
        "row_number",
        "raw_code",
        "raw_name_short",
        "detected_status_badge",
        "suggested_product_short",
        "note_short",
        "quick_actions",
    )
    list_filter = ("detected_status", "batch")
    search_fields = (
        "raw_code",
        "normalized_code",
        "raw_name",
        "suggested_product__name",
        "selected_product__name",
    )
    autocomplete_fields = ("suggested_product", "selected_product", "created_product")
    readonly_fields = (
        "batch",
        "row_number",
        "raw_code",
        "normalized_code",
        "raw_name",
        "raw_quantity",
        "quantity",
        "detected_status",
        "suggested_product",
        "note",
        "is_processed",
        "created_product",
        "inventory_updated",
        "inventory_before",
        "inventory_after",
        "created_at",
        "updated_at",
    )
    actions = ("set_action_skip", "set_action_create", "set_action_link_to_suggested")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<uuid:row_id>/quick-action/<str:action>/",
                self.admin_site.admin_view(self.quick_row_action),
                name="products_productimportrow_quick_action",
            ),
        ]
        return custom_urls + urls

    def quick_row_action(self, request, row_id, action):
        row = ProductImportRow.objects.select_related(
            "batch",
            "suggested_product",
            "selected_product",
        ).filter(id=row_id).first()

        if not row:
            messages.error(request, "Строка импорта не найдена.")
            return redirect("../..")

        try:
            if action == "skip":
                row.action = ProductImportRow.ACTION_SKIP
                row.is_processed = True
                row.save(update_fields=["action", "is_processed", "updated_at"])
                messages.success(request, f"Строка {row.row_number} пропущена.")

            elif action == "link":
                product = row.suggested_product or row.selected_product

                if not product:
                    messages.error(request, "Нет найденного похожего товара для привязки.")
                    return redirect(request.META.get("HTTP_REFERER", "../.."))

                row.selected_product = product
                row.action = ProductImportRow.ACTION_LINK
                row.is_processed = True

                if row.batch.import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
                    update_inventory_from_import_row(row=row, product=product)

                row.save(update_fields=[
                    "selected_product",
                    "action",
                    "is_processed",
                    "inventory_updated",
                    "inventory_before",
                    "inventory_after",
                    "updated_at",
                ])
                messages.success(request, f"Строка {row.row_number} привязана к существующему товару.")

            elif action == "create":
                normalized_code = normalize_product_code(row.raw_code)
                existing_product = Product.objects.filter(normalized_code=normalized_code).first()

                if existing_product:
                    row.selected_product = existing_product
                    row.action = ProductImportRow.ACTION_LINK
                    row.is_processed = True

                    if row.batch.import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
                        update_inventory_from_import_row(row=row, product=existing_product)

                    row.save(update_fields=[
                        "selected_product",
                        "action",
                        "is_processed",
                        "inventory_updated",
                        "inventory_before",
                        "inventory_after",
                        "updated_at",
                    ])
                    messages.warning(
                        request,
                        f"Товар уже найден после проверки. Строка {row.row_number} привязана к существующему товару.",
                    )
                else:
                    product = Product.objects.create(
                        name=row.raw_code.strip(),
                        description=build_import_product_description(row.raw_code, row.raw_name),
                    )

                    row.created_product = product
                    row.action = ProductImportRow.ACTION_CREATE
                    row.is_processed = True

                    if row.batch.import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
                        update_inventory_from_import_row(row=row, product=product)

                    row.save(update_fields=[
                        "created_product",
                        "action",
                        "is_processed",
                        "inventory_updated",
                        "inventory_before",
                        "inventory_after",
                        "updated_at",
                    ])
                    messages.success(request, f"Создан новый товар из строки {row.row_number}.")

            else:
                messages.error(request, "Неизвестное действие.")

        except Exception as e:
            messages.error(request, f"Ошибка обработки строки: {e}")

        return redirect(request.META.get("HTTP_REFERER", "../.."))

    def changelist_view(self, request, extra_context=None):
        queryset = self.get_queryset(request)

        batch_id = request.GET.get("batch__id__exact")
        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)

        problem_statuses = [
            ProductImportRow.STATUS_SIMILAR_CODE,
            ProductImportRow.STATUS_SIMILAR_NAME,
            ProductImportRow.STATUS_DUPLICATE_IN_FILE,
            ProductImportRow.STATUS_ERROR,
            ProductImportRow.STATUS_QUANTITY_ERROR,
        ]

        existing_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_EXISTING
        ).count()

        new_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_NEW
        ).count()

        similar_code_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_SIMILAR_CODE
        ).count()

        duplicate_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_DUPLICATE_IN_FILE
        ).count()

        quantity_error_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_QUANTITY_ERROR
        ).count()

        error_count = queryset.filter(
            detected_status=ProductImportRow.STATUS_ERROR
        ).count()

        all_error_count = quantity_error_count + error_count

        total_control_count = (
            existing_count
            + new_count
            + similar_code_count
            + duplicate_count
            + all_error_count
        )

        extra_context = extra_context or {}
        extra_context.update({
            "row_total_count": queryset.count(),
            "row_existing_count": existing_count,
            "row_new_count": new_count,
            "row_similar_code_count": similar_code_count,
            "row_duplicate_count": duplicate_count,
            "row_quantity_error_count": quantity_error_count,
            "row_error_count": error_count,
            "row_all_error_count": all_error_count,
            "row_total_control_count": total_control_count,
        })

        return super().changelist_view(request, extra_context=extra_context)


    @admin.display(description="Название")
    def raw_name_short(self, obj):
        if not obj.raw_name:
            return "—"
        if len(obj.raw_name) <= 45:
            return obj.raw_name
        return f"{obj.raw_name[:45]}..."

    @admin.display(description="Похожий товар")
    def suggested_product_short(self, obj):
        if not obj.suggested_product:
            return "—"
        return obj.suggested_product.name

    @admin.display(description="Причина")
    def note_short(self, obj):
        if not obj.note:
            return "—"
        clean_note = " ".join(str(obj.note).split())
        if len(clean_note) <= 55:
            return clean_note
        return f"{clean_note[:55]}..."

    @admin.display(description="Проверка")
    def detected_status_badge(self, obj):
        colors = {
            ProductImportRow.STATUS_EXISTING: ("#14532d", "#86efac"),
            ProductImportRow.STATUS_SIMILAR_CODE: ("#7c2d12", "#fdba74"),
            ProductImportRow.STATUS_SIMILAR_NAME: ("#713f12", "#fde68a"),
            ProductImportRow.STATUS_DUPLICATE_IN_FILE: ("#7f1d1d", "#fecaca"),
            ProductImportRow.STATUS_NEW: ("#1e3a8a", "#93c5fd"),
            ProductImportRow.STATUS_ERROR: ("#374151", "#e5e7eb"),
            ProductImportRow.STATUS_QUANTITY_ERROR: ("#7f1d1d", "#fecaca"),
        }
        bg, fg = colors.get(obj.detected_status, ("#374151", "#e5e7eb"))

        return format_html(
            '<span style="display:inline-block;min-width:135px;text-align:center;padding:6px 10px;border-radius:999px;background:{};color:{};font-weight:800;white-space:nowrap;">{}</span>',
            bg,
            fg,
            obj.get_detected_status_display(),
        )

    @admin.display(description="Действие")
    def quick_actions(self, obj):
        if obj.is_processed:
            return format_html(
                '<span style="display:inline-block;min-width:120px;text-align:center;padding:6px 10px;border-radius:999px;background:{};color:{};font-weight:800;">{}</span>',
                "#14532d",
                "#86efac",
                "Готово",
            )

        base_style = (
            "display:flex;align-items:center;justify-content:center;"
            "width:105px;margin:4px auto;padding:8px 10px;border-radius:10px;"
            "font-weight:900;text-decoration:none;color:white;"
            "box-sizing:border-box;"
        )

        skip_url = f"{obj.id}/quick-action/skip/"
        create_url = f"{obj.id}/quick-action/create/"
        link_url = f"{obj.id}/quick-action/link/"

        if obj.detected_status == ProductImportRow.STATUS_SIMILAR_CODE and obj.suggested_product:
            return format_html(
                '<a href="{}" style="{}background:#0d6efd;">Тот же</a>'
                '<a href="{}" style="{}background:#22c55e;color:#052e16;">Создать</a>'
                '<a href="{}" style="{}background:#6b7280;">Пропуск</a>',
                link_url,
                base_style,
                create_url,
                base_style,
                skip_url,
                base_style,
            )

        if obj.detected_status in [
            ProductImportRow.STATUS_DUPLICATE_IN_FILE,
            ProductImportRow.STATUS_ERROR,
            ProductImportRow.STATUS_QUANTITY_ERROR,
        ]:
            return format_html(
                '<a href="{}" style="{}background:#6b7280;">Пропуск</a>',
                skip_url,
                base_style,
            )

        return "—"

    @admin.action(description="Решение: пропустить")
    def set_action_skip(self, request, queryset):
        updated = queryset.update(action=ProductImportRow.ACTION_SKIP)
        messages.success(request, f"Обновлено строк: {updated}.")

    @admin.action(description="Решение: создать новый товар")
    def set_action_create(self, request, queryset):
        updated = queryset.update(action=ProductImportRow.ACTION_CREATE)
        messages.success(request, f"Обновлено строк: {updated}.")

    @admin.action(description="Решение: привязать к предложенному товару")
    def set_action_link_to_suggested(self, request, queryset):
        updated = 0

        for row in queryset:
            if row.suggested_product:
                row.selected_product = row.suggested_product
                row.action = ProductImportRow.ACTION_LINK
                row.save(update_fields=["selected_product", "action", "updated_at"])
                updated += 1

        messages.success(request, f"Привязано строк к предложенным товарам: {updated}.")


def update_inventory_from_import_row(row, product):
    if row.quantity is None or not row.batch.target_warehouse:
        return

    inventory, _ = Inventory.objects.get_or_create(
        product=product,
        warehouse=row.batch.target_warehouse,
        defaults={"quantity": 0},
    )

    before = inventory.quantity
    after = before + row.quantity

    inventory.quantity = after
    inventory.save(update_fields=["quantity", "updated_at"])

    row.inventory_before = before
    row.inventory_after = after
    row.inventory_updated = True


def create_product_import_batch_from_excel(
    excel_file,
    user=None,
    import_type=ProductImportBatch.IMPORT_TYPE_PRODUCTS_ONLY,
    target_warehouse=None,
    quantity_mode=ProductImportBatch.QUANTITY_MODE_REPLACE,
):
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    code_headers = [
        "name",
        "код",
        "код товара",
        "part no",
        "part number",
        "pn",
        "чертежный номер",
        "чертёжный номер",
        "наименование",
        "номенклатура",
        "обозначение",
    ]
    description_headers = [
        "description",
        "название",
        "описание",
        "item",
        "product",
        "наименование",
        "номенклатура",
        "наименование товара",
    ]
    unit_headers = [
        "unit",
        "единица",
        "ед. изм",
        "ед.изм",
        "uom",
        "единица измерения",
    ]
    quantity_headers = [
        "колич",
        "кол-во",
        "кол во",
        "количество",
        "qty",
        "quantity",
        "остаток",
        "наличие",
        "шт",
        "pcs",
    ]

    header_result = detect_excel_header_row(
        sheet=sheet,
        code_headers=code_headers,
        description_headers=description_headers,
        unit_headers=unit_headers,
        quantity_headers=quantity_headers,
        import_type=import_type,
    )

    header_row = header_result["header_row"]
    code_index = header_result["code_index"]
    description_index = header_result["description_index"]
    unit_index = header_result["unit_index"]
    quantity_index = header_result["quantity_index"]

    if code_index is None:
        raise ValueError(
            "Не нашёл колонку с чертежным номером. Проверь шаблон Excel."
        )

    if import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
        if target_warehouse is None:
            raise ValueError("Для импорта остатков нужно выбрать склад.")

        if quantity_index is None:
            raise ValueError(
                "Не нашёл колонку с количеством. Проверь шаблон Excel."
            )

    batch = ProductImportBatch.objects.create(
        title=f"Импорт товаров: {excel_file.name}",
        file=excel_file,
        import_type=import_type,
        target_warehouse=target_warehouse,
        quantity_mode=quantity_mode or ProductImportBatch.QUANTITY_MODE_ADD,
        created_by=user if user and user.is_authenticated else None,
        status=ProductImportBatch.STATUS_DRAFT,
    )

    parsed_rows = []
    normalized_code_rows = {}

    for row_number in range(header_row + 1, sheet.max_row + 1):
        raw_code = sheet.cell(row=row_number, column=code_index).value
        raw_name = sheet.cell(row=row_number, column=description_index).value if description_index else ""
        raw_unit = sheet.cell(row=row_number, column=unit_index).value if unit_index else ""
        raw_quantity = sheet.cell(row=row_number, column=quantity_index).value if quantity_index else ""

        raw_code = str(raw_code).strip() if raw_code is not None else ""
        raw_name = str(raw_name).strip() if raw_name is not None else ""
        raw_unit = str(raw_unit).strip() if raw_unit is not None else ""
        raw_quantity_text = str(raw_quantity).strip() if raw_quantity is not None else ""

        if not raw_code and not raw_name and not raw_quantity_text:
            continue

        if looks_like_section_title(raw_code, raw_name, raw_quantity_text):
            continue

        normalized_code = normalize_product_code(raw_code)

        parsed_rows.append({
            "row_number": row_number,
            "raw_code": raw_code,
            "raw_name": raw_name,
            "raw_unit": raw_unit,
            "raw_quantity_text": raw_quantity_text,
            "raw_quantity": raw_quantity,
            "normalized_code": normalized_code,
        })

        if normalized_code:
            normalized_code_rows.setdefault(normalized_code, []).append(row_number)

    duplicate_normalized_codes = {
        code
        for code, row_numbers in normalized_code_rows.items()
        if len(row_numbers) > 1
    }

    for row_data in parsed_rows:
        row_number = row_data["row_number"]
        raw_code = row_data["raw_code"]
        raw_name = row_data["raw_name"]
        raw_unit = row_data["raw_unit"]
        raw_quantity_text = row_data["raw_quantity_text"]
        raw_quantity = row_data["raw_quantity"]
        normalized_code = row_data["normalized_code"]

        suggested_product, detected_status, note = find_best_product_match(raw_code, raw_name)

        quantity = None

        if import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
            quantity, quantity_note = parse_quantity(raw_quantity)
            if quantity_note:
                detected_status = ProductImportRow.STATUS_QUANTITY_ERROR
                note = f"{note}\n{quantity_note}"

        if normalized_code and normalized_code in duplicate_normalized_codes:
            rows_text = ", ".join(str(number) for number in normalized_code_rows[normalized_code])
            detected_status = ProductImportRow.STATUS_DUPLICATE_IN_FILE
            suggested_product = None
            note = f"Дубль внутри Excel. Этот код встречается в строках: {rows_text}."

        action = ProductImportRow.ACTION_SKIP

        if detected_status == ProductImportRow.STATUS_NEW:
            action = ProductImportRow.ACTION_CREATE
        elif detected_status == ProductImportRow.STATUS_EXISTING:
            if import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
                action = ProductImportRow.ACTION_LINK
            else:
                action = ProductImportRow.ACTION_SKIP
        elif detected_status in [
            ProductImportRow.STATUS_SIMILAR_CODE,
            ProductImportRow.STATUS_SIMILAR_NAME,
            ProductImportRow.STATUS_DUPLICATE_IN_FILE,
            ProductImportRow.STATUS_ERROR,
            ProductImportRow.STATUS_QUANTITY_ERROR,
        ]:
            action = ProductImportRow.ACTION_SKIP

        ProductImportRow.objects.create(
            batch=batch,
            row_number=row_number,
            raw_code=raw_code,
            raw_name=raw_name,
            raw_unit=raw_unit,
            raw_quantity=raw_quantity_text,
            quantity=quantity,
            normalized_code=normalized_code,
            detected_status=detected_status,
            suggested_product=suggested_product,
            selected_product=suggested_product if detected_status == ProductImportRow.STATUS_EXISTING else None,
            action=action,
            note=note.strip(),
        )

    recheck_product_import_batch(batch)
    return batch


def normalize_header_value(value):
    if value is None:
        return ""

    value = str(value).strip().lower()
    value = value.replace("\n", " ")
    value = value.replace("\r", " ")
    value = " ".join(value.split())

    return value


def find_header_index(headers, possible_names):
    normalized_possible_names = [
        normalize_header_value(name)
        for name in possible_names
        if normalize_header_value(name)
    ]

    for possible_name in normalized_possible_names:
        if possible_name in headers:
            return headers.index(possible_name) + 1

    for index, header in enumerate(headers):
        for possible_name in normalized_possible_names:
            if possible_name and possible_name in header:
                return index + 1

    return None


def detect_excel_header_row(
    sheet,
    code_headers,
    description_headers,
    unit_headers,
    quantity_headers,
    import_type,
    max_scan_rows=25,
):
    """
    Ищет строку заголовков не только в первой строке, а в первых строках Excel.

    Это нужно для реальных складских файлов, где сверху часто бывают:
    пустые строки, названия склада, даты, категории, шапки с объединёнными ячейками.
    """
    best_result = None
    best_score = -1

    scan_limit = min(sheet.max_row, max_scan_rows)

    for row_number in range(1, scan_limit + 1):
        headers = [
            normalize_header_value(cell.value)
            for cell in sheet[row_number]
        ]

        code_index = find_header_index(headers, code_headers)
        description_index = find_header_index(headers, description_headers)
        unit_index = find_header_index(headers, unit_headers)
        quantity_index = find_header_index(headers, quantity_headers)

        score = 0

        if code_index is not None:
            score += 10

        if description_index is not None:
            score += 3

        if unit_index is not None:
            score += 1

        if quantity_index is not None:
            score += 8

        if (
            import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK
            and code_index is not None
            and quantity_index is not None
        ):
            score += 20

        if (
            import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_ONLY
            and code_index is not None
        ):
            score += 15

        if score > best_score:
            best_score = score
            best_result = {
                "header_row": row_number,
                "headers": headers,
                "code_index": code_index,
                "description_index": description_index,
                "unit_index": unit_index,
                "quantity_index": quantity_index,
            }

    if best_result is None:
        best_result = {
            "header_row": 1,
            "headers": [],
            "code_index": None,
            "description_index": None,
            "unit_index": None,
            "quantity_index": None,
        }

    return best_result


def looks_like_section_title(raw_code, raw_name, raw_quantity):
    """
    Пропускает строки-разделы типа:
    ПОДШИПНИКИ
    ЭЛЕКТРООБОРУДОВАНИЕ
    КРЕПЕЖ
    если рядом нет количества и нормального кода.
    """
    text = str(raw_code or raw_name or "").strip()

    if not text:
        return False

    if raw_quantity not in [None, ""]:
        return False

    normalized = normalize_product_code(text)

    if not normalized:
        return False

    has_digit = any(char.isdigit() for char in normalized)
    has_letter = any(char.isalpha() for char in normalized)

    # Обычно реальные авиационные коды имеют цифры.
    # А строки-разделы часто только буквами.
    if has_letter and not has_digit and len(normalized) >= 4:
        return True

    return False


def recheck_product_import_batch(batch):
    rows = batch.rows.all().order_by("row_number")
    unprocessed_rows = rows.filter(is_processed=False)

    total_rows = rows.count()

    existing_count = rows.filter(
        detected_status=ProductImportRow.STATUS_EXISTING
    ).count()

    new_count = rows.filter(
        detected_status=ProductImportRow.STATUS_NEW
    ).count()

    similar_code_count = unprocessed_rows.filter(
        detected_status=ProductImportRow.STATUS_SIMILAR_CODE
    ).count()

    similar_name_count = unprocessed_rows.filter(
        detected_status=ProductImportRow.STATUS_SIMILAR_NAME
    ).count()

    duplicate_in_file_count = unprocessed_rows.filter(
        detected_status=ProductImportRow.STATUS_DUPLICATE_IN_FILE
    ).count()

    error_count = unprocessed_rows.filter(
        detected_status=ProductImportRow.STATUS_ERROR
    ).count()

    quantity_error_count = unprocessed_rows.filter(
        detected_status=ProductImportRow.STATUS_QUANTITY_ERROR
    ).count()

    batch.total_rows = total_rows
    batch.existing_count = existing_count
    batch.similar_code_count = similar_code_count
    batch.similar_name_count = similar_name_count
    batch.duplicate_in_file_count = duplicate_in_file_count
    batch.new_count = new_count
    batch.error_count = error_count
    batch.quantity_error_count = quantity_error_count

    if batch.status != ProductImportBatch.STATUS_IMPORTED:
        batch.status = ProductImportBatch.STATUS_CHECKED

    batch.save(update_fields=[
        "total_rows",
        "existing_count",
        "similar_code_count",
        "similar_name_count",
        "duplicate_in_file_count",
        "new_count",
        "error_count",
        "quantity_error_count",
        "status",
        "updated_at",
    ])


@transaction.atomic
def confirm_product_import_batch(batch):
    if batch.status == ProductImportBatch.STATUS_IMPORTED:
        return {"created": 0, "skipped": 0, "linked": 0, "inventory_updated": 0}

    created = 0
    skipped = 0
    linked = 0
    inventory_updated = 0

    rows = batch.rows.select_related("selected_product", "suggested_product").all()

    for row in rows:
        if row.is_processed:
            continue

        product = None

        if row.action == ProductImportRow.ACTION_SKIP:
            problem_statuses = [
                ProductImportRow.STATUS_SIMILAR_CODE,
                ProductImportRow.STATUS_SIMILAR_NAME,
                ProductImportRow.STATUS_DUPLICATE_IN_FILE,
                ProductImportRow.STATUS_ERROR,
                ProductImportRow.STATUS_QUANTITY_ERROR,
            ]

            if row.detected_status in problem_statuses:
                row.is_processed = False
                row.save(update_fields=["is_processed", "updated_at"])
            else:
                row.is_processed = True
                row.save(update_fields=["is_processed", "updated_at"])

            skipped += 1
            continue

        if row.action == ProductImportRow.ACTION_LINK:
            product = row.selected_product or row.suggested_product

            if product:
                row.selected_product = product
                linked += 1
            else:
                row.action = ProductImportRow.ACTION_SKIP
                row.note = f"{row.note}\nНе выбран товар для привязки."
                row.is_processed = True
                row.save(update_fields=["action", "note", "is_processed", "updated_at"])
                skipped += 1
                continue

        if row.action == ProductImportRow.ACTION_CREATE:
            normalized_code = normalize_product_code(row.raw_code)
            existing_product = Product.objects.filter(normalized_code=normalized_code).first()

            if existing_product:
                product = existing_product
                row.selected_product = existing_product
                row.action = ProductImportRow.ACTION_LINK
                linked += 1
            else:
                product = Product.objects.create(
                    name=row.raw_code.strip(),
                    description=build_import_product_description(row.raw_code, row.raw_name),
                )
                row.created_product = product
                created += 1

        if batch.import_type == ProductImportBatch.IMPORT_TYPE_PRODUCTS_WITH_STOCK:
            if product and row.quantity is not None and batch.target_warehouse:
                inventory, _ = Inventory.objects.get_or_create(
                    product=product,
                    warehouse=batch.target_warehouse,
                    defaults={"quantity": 0},
                )

                before = inventory.quantity

                if batch.quantity_mode == ProductImportBatch.QUANTITY_MODE_ADD:
                    after = before + row.quantity
                else:
                    after = row.quantity

                inventory.quantity = after
                inventory.save(update_fields=["quantity", "updated_at"])

                row.inventory_before = before
                row.inventory_after = after
                row.inventory_updated = True
                inventory_updated += 1

        row.is_processed = True
        row.save(update_fields=[
            "selected_product",
            "action",
            "is_processed",
            "created_product",
            "inventory_updated",
            "inventory_before",
            "inventory_after",
            "updated_at",
        ])

    batch.status = ProductImportBatch.STATUS_IMPORTED
    batch.created_products_count = created
    batch.updated_inventory_count = inventory_updated
    batch.skipped_rows_count = skipped
    batch.linked_rows_count = linked
    batch.save(update_fields=[
        "status",
        "created_products_count",
        "updated_inventory_count",
        "skipped_rows_count",
        "linked_rows_count",
        "updated_at",
    ])

    return {
        "created": created,
        "skipped": skipped,
        "linked": linked,
        "inventory_updated": inventory_updated,
    }
